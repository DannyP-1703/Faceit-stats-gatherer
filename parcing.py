from parser_funcs import refresh_stats
from os.path import exists
from openpyxl import Workbook, load_workbook

USERNAMES = ('nonyl', 's1mple')
TABLE_NAME = "FaceIt Stats.xlsx"

if not exists(TABLE_NAME):
    Workbook().save(filename=TABLE_NAME)
wb = load_workbook(TABLE_NAME)

for i, un in enumerate(USERNAMES):
    refresh_stats(wb, un, i)
wb.save(TABLE_NAME)
