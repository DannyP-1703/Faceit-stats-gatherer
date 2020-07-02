import sys
from datetime import datetime
from openpyxl.formatting import Rule
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter as gcl

DATE_FORMAT_IN = "%Y-%m-%d %H:%M"
DATE_FORMAT_OUT = "%d.%m.%Y %H:%M"

#       Colours
white = 'FFFFFF'
maroon = '800000'
purple = 'CCC0DA'
win = '90EE90'
lose = 'FFB6C1'
red = 'FFA07A'
yellow = 'FFFF00'
green = '32CD32'
map_colors = {'de_mirage': 'FFFF00',
              'de_inferno': '4F6228',
              'de_dust2': 'F79646',
              'de_overpass': 'C4D79B',
              'de_nuke': '8DB4E2',
              'de_vertigo': 'A6A6A6',
              'de_train': 'C4BD97',
              'de_cache': 'FFFFFF'
              }


def calc_last_row(xl_ws):       # the number of row, after the last non-empty row
    last_row = 2
    while xl_ws[f'A{last_row}'].value is not None:
        last_row += 1
    return last_row


def strin_to_date(date_string):
    return datetime.strptime(date_string, DATE_FORMAT_IN)


def strout_to_date(date_string):
    return datetime.strptime(date_string, DATE_FORMAT_OUT)


def add_template(xl_ws):
    #       Text and formulas
    xl_ws['A1'] = 'Result'
    xl_ws['B1'] = 'Team'
    xl_ws['C1'] = 'K/A/D'
    xl_ws['D1'] = 'K/R'
    xl_ws['E1'] = 'K/D'
    xl_ws['F1'] = 'HS rate'
    xl_ws['G1'] = 'Score'
    xl_ws['H1'] = 'Map'
    xl_ws['I1'] = 'Date'
    xl_ws['J1'] = 'ELO diff.'
    xl_ws['K1'] = 'ELO aft. match'
    xl_ws['N2'] = 'Total games'
    xl_ws['N3'] = '=COUNTA(A:A)-1'
    xl_ws['O2'] = 'Win rate'
    xl_ws['O3'] = '=COUNTIF(A:A,"Win")/N3'
    xl_ws['P2'] = 'Average K/D'
    xl_ws['P3'] = '=AVERAGE(E:E)'
    xl_ws['Q2'] = 'Av. HS rate'
    xl_ws['Q3'] = '=AVERAGE(F:F)'
    xl_ws['N5'] = 'Map'
    xl_ws['N6'] = 'Map %'
    xl_ws['N7'] = 'Win rate %'
    xl_ws['O5'] = 'de_mirage'
    xl_ws['P5'] = 'de_inferno'
    xl_ws['Q5'] = 'de_dust2'
    xl_ws['R5'] = 'de_overpass'
    xl_ws['S5'] = 'de_nuke'
    xl_ws['T5'] = 'de_vertigo'
    xl_ws['U5'] = 'de_train'
    for i in range(15, 22):
        col_letter = gcl(i)
        xl_ws.cell(row=6, column=i).value = f'=COUNTIF(H:H,{col_letter}5)/N3'
        xl_ws.cell(row=7, column=i).value = f'=IF(COUNTIF(H:H,{col_letter}5)=0,0,' \
                                            f'COUNTIFS(H:H,"="&{col_letter}5,A:A,"=Win")/COUNTIF(H:H,{col_letter}5))'

        # Conditional formatting
        win_rule = Rule(type="containsText", operator="containsText", text="Win",
                        dxf=DifferentialStyle(fill=PatternFill(patternType='solid', bgColor=win)))
        win_rule.formula = ['NOT(ISERROR(SEARCH("Win",A1)))']
        xl_ws.conditional_formatting.add('A1:A1048576', win_rule)

        lose_rule = Rule(type="containsText",
                         operator="containsText", text="Lose",
                         dxf=DifferentialStyle(fill=PatternFill(patternType='solid', bgColor=lose)))
        lose_rule.formula = ['NOT(ISERROR(SEARCH("Lose",A1)))']
        xl_ws.conditional_formatting.add('A1:A1048576', lose_rule)

        colour_scale_rule = ColorScaleRule(start_type='min', start_color=red,
                                           mid_type='percentile', mid_color=yellow,
                                           end_type='max', end_color=green)

        xl_ws.conditional_formatting.add('E1:E1048576', colour_scale_rule)
        xl_ws.conditional_formatting.add('F1:F1048576', colour_scale_rule)


def fill_in_stats(xl_ws, stats):
    last_row = calc_last_row(xl_ws)

    last_date = datetime(year=1, month=1, day=1, hour=0, minute=0)
    if last_row > 2:
        last_date = strout_to_date(xl_ws[f'I{last_row - 1}'].value)

    for m in stats:
        if last_date >= strin_to_date(m[8]):
            continue
        xl_ws.cell(row=last_row, column=1).value = m[0]
        xl_ws.cell(row=last_row, column=2).value = m[1]
        xl_ws.cell(row=last_row, column=3).value = m[2]
        xl_ws.cell(row=last_row, column=4).value = float(m[3])
        xl_ws.cell(row=last_row, column=5).value = float(m[4])
        xl_ws.cell(row=last_row, column=6).value = float(m[5][:m[5].find('%')]) / 100
        xl_ws.cell(row=last_row, column=7).value = m[6]
        xl_ws.cell(row=last_row, column=8).value = m[7]
        xl_ws.cell(row=last_row, column=9).value = strin_to_date(m[8]).strftime(DATE_FORMAT_OUT)
        try:
            xl_ws.cell(row=last_row, column=10).value = int(m[10])
            xl_ws.cell(row=last_row, column=11).value = int(m[9])
        except IndexError:
            sys.stderr.write("Some ELO stats are missing in the site\n")
        last_row += 1


def apply_styles(xl_ws):
    #   Static formatting
    last_row = calc_last_row(xl_ws)

    for i in range(1, 23):
        if i == 4 or i == 5 or i == 6:
            xl_ws.column_dimensions[gcl(i)].width = 8.86
        elif i == 2 or i == 9:
            xl_ws.column_dimensions[gcl(i)].width = 16
        else:
            xl_ws.column_dimensions[gcl(i)].width = 12.14
    for i in range(1, last_row):
        xl_ws.row_dimensions[i].height = 18.75

    for row in xl_ws[f'A2:K{last_row - 1}']:
        for c in row:
            c.alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)
            c.border = Border(right=Side(border_style='thin'))
            c.font = Font(size=12)

    for row in xl_ws['A1:K1']:
        for c in row:
            c.font = Font(size=14, bold=True)
            c.alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)
            c.border = Border(bottom=Side(border_style='thick'), right=Side(border_style='thin'))
            c.fill = PatternFill('solid', fgColor=purple)

    all_borders = Border(top=Side(border_style='thin'),
                         bottom=Side(border_style='thin'),
                         left=Side(border_style='thin'),
                         right=Side(border_style='thin'))

    for row in xl_ws['N2:Q3']:
        for c in row:
            c.alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)
            c.font = Font(size=14)
            c.border = all_borders

    for row in xl_ws['N5:U7']:
        for c in row:
            c.alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)
            c.font = Font(size=14)
            c.border = all_borders

    for row in xl_ws['O6:U7']:
        for c in row:
            c.font = Font(size=14, color=maroon)
            c.number_format = '0%'

    for i in range(14, 18):
        xl_ws[f'{gcl(i)}2'].fill = PatternFill('solid', fgColor=purple)

    for i in range(5, 8):
        xl_ws[f'N{i}'].fill = PatternFill('solid', fgColor=purple)

    xl_ws['O3'].number_format = '0.00%'
    xl_ws['P3'].number_format = '0.00'
    xl_ws['Q3'].number_format = '0.0%'
    for i in range(1, last_row):
        xl_ws[f'D{i}'].number_format = '0.00'
        xl_ws[f'E{i}'].number_format = '0.00'
        xl_ws[f'F{i}'].number_format = '0%'
        xl_ws[f'J{i}'].number_format = '+#;-#'

    for i in range(15, 22):
        col_letter = gcl(i)
        c = xl_ws[f'{col_letter}5']
        c.fill = PatternFill('solid', fgColor=map_colors[c.value])
        if i == 16:
            c.font = Font(size=14, color=white)

    for row in xl_ws.iter_rows(min_row=2, min_col=8, max_col=8, max_row=last_row - 1):
        for c in row:
            c.fill = PatternFill(patternType='solid', fgColor=map_colors[c.value])
            if c.value == 'de_inferno':
                c.font = Font(size=12, color=white)
